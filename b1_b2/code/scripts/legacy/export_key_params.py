#!/usr/bin/env python3
"""
OceanBase v4.5 关键性能参数导出脚本

从GV$OB_PARAMETERS中筛选100个关键性能参数，按4个维度分类：
- CPU调度
- 内存管理
- 磁盘IO
- SQL执行

标注每个参数：默认值、取值范围、核心作用，导出Excel
"""

import pymysql
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

# ============================================================
# 100个关键性能参数定义
# ============================================================
# 格式: (参数名, 维度, 默认值, 取值范围, 核心作用)
KEY_PARAMS = [
    # ======================== CPU调度 (25个) ========================
    ("cpu_count", "CPU调度", "0", "[0, +∞) 整数，0=自动检测", "系统CPU核数，0表示自动检测。直接影响线程池和调度器规模"),
    ("cpu_quota_concurrency", "CPU调度", "4", "[1, 20]", "单个CPU配额允许的最大并发数，控制租户并发上限"),
    ("workers_per_cpu_quota", "CPU调度", "10", "[1, +∞) 整数", "每个CPU配额分配的工作线程数，决定租户线程池大小"),
    ("large_query_worker_percentage", "CPU调度", "30", "[0, 100] 百分比", "为大查询预留的工作线程百分比，防止大查询饿死小查询"),
    ("large_query_threshold", "CPU调度", "5s", "[0s, +∞)", "大查询判定阈值，超过此时间的请求可能被降级处理"),
    ("_ob_enable_dynamic_worker", "CPU调度", "True", "True/False", "是否启用动态工作线程扩展，所有worker忙碌时自动扩容"),
    ("_stall_threshold_for_dynamic_worker", "CPU调度", "3ms", "[1ms, +∞)", "动态worker触发阈值，请求排队超过此时间则扩容worker"),
    ("_enable_more_aggressive_dynamic_worker", "CPU调度", "True", "True/False", "更激进地增加worker数量，在高并发场景下更快扩展"),
    ("global_background_cpu_quota", "CPU调度", "-1", "[-1, +∞)", "后台任务CPU配额，-1表示不隔离前后台资源"),
    ("enable_global_background_resource_isolation", "CPU调度", "False", "True/False", "前后台资源隔离开关，True时使用独立后台CPU配额"),
    ("server_cpu_quota_max", "CPU调度", "0", "[0, +∞)", "server租户最大vCPU配额，0表示不限制"),
    ("server_cpu_quota_min", "CPU调度", "0", "[0, +∞)", "server租户最小vCPU配额，保障系统租户最低CPU资源"),
    ("net_thread_count", "CPU调度", "0", "[0, +∞) 整数", "RPC/MySQL IO线程数，0表示自动计算。影响网络请求处理能力"),
    ("high_priority_net_thread_count", "CPU调度", "0", "[0, +∞) 整数", "高优先级RPC IO线程数，0表示不单独分配"),
    ("sql_net_thread_count", "CPU调度", "0", "[0, 64] 整数", "全局MySQL IO线程数，0表示自动计算"),
    ("tenant_sql_net_thread_count", "CPU调度", "0", "[0, 64] 整数", "每个租户的MySQL IO线程数，0表示自动计算"),
    ("sql_login_thread_count", "CPU调度", "0", "[0, 32] 整数", "SQL登录请求处理线程数，0表示自动计算"),
    ("tenant_sql_login_thread_count", "CPU调度", "0", "[0, +∞) 整数", "每个租户SQL登录线程数"),
    ("io_scheduler_thread_count", "CPU调度", "4", "[1, 16] 整数", "IO调度线程数，控制IO请求分发能力"),
    ("disk_io_thread_count", "CPU调度", "8", "[1, +∞) 整数", "每块磁盘的IO线程数，影响磁盘IO处理并发度"),
    ("sync_io_thread_count", "CPU调度", "0", "[0, +∞) 整数", "同步IO线程数，0表示自动配置"),
    ("_io_callback_thread_count", "CPU调度", "0", "[0, +∞) 整数", "IO回调线程数，0表示自动配置。影响IO完成通知处理"),
    ("_ob_max_thread_num", "CPU调度", "0", "[0, +∞) 整数", "observer线程数上限，0表示不限制。防止线程数爆炸"),
    ("tenant_task_queue_size", "CPU调度", "16384", "[1024, +∞)", "每个租户的任务队列大小，影响请求排队容量"),
    ("location_cache_cpu_quota", "CPU调度", "5", "[0, +∞)", "位置缓存请求分配的vCPU数，影响路由信息更新速度"),

    # ======================== 内存管理 (25个) ========================
    ("memory_limit", "内存管理", "0M", "[0M, +∞)", "observer可用内存总量，0M表示用memory_limit_percentage计算"),
    ("memory_limit_percentage", "内存管理", "80", "[10, 95] 百分比", "observer内存占总物理内存的百分比，关键内存上限参数"),
    ("system_memory", "内存管理", "0M", "[0M, +∞)", "系统内部预留内存，不可被租户分配使用"),
    ("memstore_limit_percentage", "内存管理", "0", "[1, 99] 百分比", "MemStore占租户内存的百分比，控制写入内存上限"),
    ("freeze_trigger_percentage", "内存管理", "20", "[1, 99] 百分比", "MemStore使用率达到此阈值触发冻结，防止内存溢出"),
    ("query_memory_limit_percentage", "内存管理", "50", "[0, 100] 百分比", "单个查询可使用的租户内存百分比上限，防止单查询耗尽内存"),
    ("sql_work_area", "内存管理", "1G", "[1M, +∞)", "租户SQL工作区内存限制，影响排序/哈希等操作内存"),
    ("_hash_area_size", "内存管理", "32M", "[2M, +∞)", "HASH JOIN可使用的最大内存"),
    ("_sort_area_size", "内存管理", "32M", "[2M, +∞)", "排序操作可使用的最大内存"),
    ("cache_wash_threshold", "内存管理", "4GB", "[0M, +∞)", "缓存淘汰触发阈值，剩余内存低于此值时开始淘汰缓存"),
    ("memory_chunk_cache_size", "内存管理", "0M", "[0M, +∞)", "内存块缓存最大值，0表示自动管理。影响内存分配效率"),
    ("memory_reserved", "内存管理", "500M", "[0M, +∞)", "系统预留的紧急内存，用于极端情况下的内部操作"),
    ("_storage_meta_memory_limit_percentage", "内存管理", "20", "[0, 100] 百分比", "存储元数据占租户内存的百分比上限"),
    ("_mds_memory_limit_percentage", "内存管理", "10", "[0, 100] 百分比", "MDS（多版本数据服务）占租户内存的百分比上限"),
    ("_tx_data_memory_limit_percentage", "内存管理", "20", "[0, 100] 百分比", "事务数据占租户内存的百分比上限"),
    ("_tx_share_memory_limit_percentage", "内存管理", "0", "[0, 100] 百分比", "事务共享内存占租户内存的百分比上限"),
    ("rpc_memory_limit_percentage", "内存管理", "0", "[0, 100] 百分比", "RPC占用租户内存的百分比上限，0表示不限制"),
    ("ob_vector_memory_limit_percentage", "内存管理", "0", "[0, 100] 百分比", "向量索引占用租户内存的百分比上限"),
    ("_ctx_memory_limit", "内存管理", "", "字符串格式，如'1G'", "租户上下文内存限制，控制会话级内存使用"),
    ("_chunk_row_store_mem_limit", "内存管理", "0M", "[0M, +∞)", "ChunkRowStore最大内存，0表示跟随sql_work_area"),
    ("result_cache_max_size", "内存管理", "64M", "[0M, +∞)", "结果缓存最大内存，影响查询结果缓存容量"),
    ("result_cache_max_result", "内存管理", "5", "[0, 100] 百分比", "单条结果缓存占结果缓存总量的百分比上限"),
    ("_px_max_message_pool_pct", "内存管理", "40", "[0, 100] 百分比", "DTL消息缓冲池占租户内存的百分比上限"),
    ("_parallel_min_message_pool", "内存管理", "16M", "[1M, +∞)", "DTL消息缓冲池保留的最小内存"),
    ("_temporary_file_io_area_size", "内存管理", "1", "[0, 100] 百分比", "临时文件IO缓冲区占租户内存的百分比"),

    # ======================== 磁盘IO (25个) ========================
    ("datafile_disk_percentage", "磁盘IO", "0", "[0, 99] 百分比", "数据文件占用磁盘空间的百分比，0表示使用datafile_size"),
    ("data_disk_usage_limit_percentage", "磁盘IO", "90", "[50, 100] 百分比", "数据磁盘安全使用率百分比，超限触发告警"),
    ("data_disk_write_limit_percentage", "磁盘IO", "0", "[0, 100] 百分比", "数据磁盘写入限制百分比，超限停止用户写入"),
    ("log_disk_percentage", "磁盘IO", "0", "[0, 99] 百分比", "日志文件占用磁盘空间的百分比，0表示使用log_disk_size"),
    ("log_disk_size", "磁盘IO", "0M", "[0M, +∞)", "日志文件磁盘空间大小"),
    ("log_disk_throttling_percentage", "磁盘IO", "60", "[0, 100] 百分比", "日志磁盘限流阈值百分比，超限后写入限速"),
    ("log_disk_throttling_maximum_duration", "磁盘IO", "2h", "[1s, +∞)", "日志磁盘限流最大持续时间"),
    ("log_disk_utilization_threshold", "磁盘IO", "80", "[0, 100] 百分比", "日志磁盘利用率阈值，超限后重用日志文件"),
    ("log_disk_utilization_limit_threshold", "磁盘IO", "95", "[0, 100] 百分比", "日志磁盘利用率上限，超限停止提交日志"),
    ("_data_storage_io_timeout", "磁盘IO", "10s", "[1s, 600s]", "数据存储IO超时时间，影响IO等待上限"),
    ("_io_read_batch_size", "磁盘IO", "0K", "[0K, 16M]", "单次读IO请求最大批量大小，0表示使用默认值"),
    ("_io_read_redundant_limit_percentage", "磁盘IO", "0", "[0, 100] 百分比", "单次读IO请求冗余数据百分比上限"),
    ("clog_io_isolation_mode", "磁盘IO", "1", "[1, 2] 整数", "CLog IO隔离模式，1=共享模式，2=独占模式"),
    ("clog_sync_time_warn_threshold", "磁盘IO", "100ms", "[1ms, +∞)", "CLog同步时间告警阈值，超限记录告警日志"),
    ("_log_writer_parallelism", "磁盘IO", "3", "[1, +∞) 整数", "并行日志写入线程数，影响日志写入吞吐"),
    ("_enable_parallel_redo_logging", "磁盘IO", "True", "True/False", "是否启用并行REDO日志写入，提升日志写入性能"),
    ("_parallel_redo_logging_trigger", "磁盘IO", "16M", "[1M, +∞)", "触发并行REDO日志写入的待写日志大小阈值"),
    ("_enable_tree_based_io_scheduler", "磁盘IO", "True", "True/False", "是否启用基于树的IO调度器，优化IO调度策略"),
    ("ss_cache_max_percentage", "磁盘IO", "30", "[0, 100] 百分比", "本地缓存磁盘空间占数据文件空间的最大百分比"),
    ("ss_cache_maxsize_percpu", "磁盘IO", "128G", "[0M, +∞)", "每CPU允许的最大本地缓存磁盘空间"),
    ("_ss_micro_cache_memory_percentage", "磁盘IO", "20", "[0, 100] 百分比", "微块缓存占租户内存的百分比"),
    ("_ss_micro_cache_size_max_percentage", "磁盘IO", "20", "[0, 100] 百分比", "微块缓存占租户磁盘空间的最大百分比"),
    ("writing_throttling_maximum_duration", "磁盘IO", "2h", "[1s, +∞)", "写入限流最大持续时间，防止长时间写入阻塞"),
    ("_private_buffer_size", "磁盘IO", "16K", "[0B, +∞)", "事务私有缓冲区触发大小，影响事务日志批量写入"),
    ("_object_storage_io_timeout", "磁盘IO", "20s", "[1s, 1200s]", "对象存储IO超时时间"),

    # ======================== SQL执行 (25个) ========================
    ("enable_sql_audit", "SQL执行", "True", "True/False", "SQL审计开关，记录SQL执行统计信息，性能诊断基础"),
    ("trace_log_slow_query_watermark", "SQL执行", "1s", "[0s, +∞)", "慢查询日志阈值，超过此时间的查询记录到慢日志"),
    ("plan_cache_evict_interval", "SQL执行", "5s", "[0s, +∞)", "执行计划缓存淘汰间隔，控制计划缓存刷新频率"),
    ("enable_adaptive_plan_cache", "SQL执行", "False", "True/False", "自适应计划缓存，根据执行统计自动优化计划选择"),
    ("_pc_adaptive_effectiveness_ratio_threshold", "SQL执行", "5", "[1, +∞) 整数", "自适应计划缓存最低有效性比率阈值"),
    ("_ob_plan_cache_auto_flush_interval", "SQL执行", "0s", "[0s, +∞)", "计划缓存自动刷新间隔，0表示不自动刷新"),
    ("_ob_plan_cache_gc_strategy", "SQL执行", "REPORT", "OFF/REPORT/RELEASE", "计划缓存GC策略，OFF=关闭，REPORT=仅报告，RELEASE=释放"),
    ("enable_sql_operator_dump", "SQL执行", "True", "True/False", "SQL算子(排序/哈希/物化)是否允许磁盘溢写，防止内存不足"),
    ("spill_compression_codec", "SQL执行", "NONE", "NONE/ZSTD/LZ4等", "磁盘溢写压缩算法，减少溢写IO量"),
    ("_force_hash_join_spill", "SQL执行", "False", "True/False", "强制哈希连接溢写，用于调试和测试"),
    ("_force_hash_groupby_dump", "SQL执行", "False", "True/False", "强制哈希聚合溢写，用于调试和测试"),
    ("_hash_join_enabled", "SQL执行", "True", "True/False", "是否启用哈希连接，关闭后优化器不选择Hash Join"),
    ("_nested_loop_join_enabled", "SQL执行", "True", "True/False", "是否启用嵌套循环连接"),
    ("_optimizer_sortmerge_join_enabled", "SQL执行", "True", "True/False", "是否启用归并排序连接"),
    ("_enable_hash_join_hasher", "SQL执行", "1", "[1, 2] 整数，1=murmurhash, 2=crc64", "哈希连接使用的哈希函数选择"),
    ("_enable_hash_join_processor", "SQL执行", "7", "[1, 7] 整数", "哈希连接处理路径选择，7=自动选择"),
    ("_bloom_filter_ratio", "SQL执行", "35", "[0, 100]", "PX布隆过滤器误判率，影响并行查询过滤效率"),
    ("px_workers_per_cpu_quota", "SQL执行", "10", "[1, +∞) 整数", "每个CPU配额分配的PX工作线程数"),
    ("_max_px_workers_per_cpu", "SQL执行", "1", "[1, +∞) 整数", "每CPU允许的PX工作线程上限"),
    ("_parallel_max_active_sessions", "SQL执行", "0", "[0, +∞) 整数", "租户最大活跃并行会话数，0=不限制"),
    ("_px_max_pipeline_depth", "SQL执行", "2", "[2, 3] 整数", "并行执行最大流水线深度"),
    ("_px_join_skew_handling", "SQL执行", "True", "True/False", "并行连接数据倾斜处理开关，优化倾斜场景"),
    ("_px_join_skew_minfreq", "SQL执行", "30", "[1, 100] 百分比", "并行连接倾斜值最小频率阈值"),
    ("_enable_index_merge", "SQL执行", "False", "True/False", "索引合并优化开关，可合并多个索引扫描结果"),
    ("optimizer_index_cost_adj", "SQL执行", "0", "[0, +∞) 整数", "索引扫描成本调整因子，影响优化器选择索引的倾向"),
]


def fetch_current_values():
    """从数据库获取当前参数值"""
    try:
        conn = pymysql.connect(
            host=os.getenv("OB_HOST", "127.0.0.1"),
            port=int(os.getenv("OB_PORT", "2881")),
            user=os.getenv("OB_USER", "root@sys"),
            password=os.getenv("OB_PASSWORD", ""),
            database=os.getenv("OB_DATABASE", "oceanbase"),
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor
        )
        with conn.cursor() as cur:
            cur.execute("SELECT NAME, VALUE, DEFAULT_VALUE, DATA_TYPE, EDIT_LEVEL, SECTION, INFO FROM oceanbase.GV$OB_PARAMETERS")
            rows = cur.fetchall()
            param_map = {}
            for row in rows:
                param_map[row['NAME']] = {
                    'current_value': row['VALUE'],
                    'default_value': row['DEFAULT_VALUE'],
                    'data_type': row['DATA_TYPE'],
                    'edit_level': row['EDIT_LEVEL'],
                    'section': row['SECTION'],
                    'info': row['INFO'] or ''
                }
        conn.close()
        return param_map
    except Exception as e:
        print(f"⚠ 无法连接数据库获取当前值: {e}")
        return {}


def export_to_excel(params, param_map, output_path):
    """导出参数到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OceanBase关键性能参数"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cpu_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    mem_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    io_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    sql_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    dim_fills = {
        "CPU调度": cpu_fill,
        "内存管理": mem_fill,
        "磁盘IO": io_fill,
        "SQL执行": sql_fill,
    }

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    # 表头
    headers = ["序号", "维度", "参数名", "当前值", "默认值", "取值范围", "数据类型", "编辑级别", "核心作用"]
    col_widths = [6, 10, 42, 18, 18, 30, 10, 18, 65]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 数据行
    for idx, (name, dim, default_val, value_range, core_func) in enumerate(params, 1):
        row_num = idx + 1
        db_info = param_map.get(name, {})

        # 如果数据库中有实际值，优先使用
        current_val = db_info.get('current_value', '-')
        db_default = db_info.get('default_value', default_val)
        data_type = db_info.get('data_type', '-')
        edit_level = db_info.get('edit_level', '-')

        row_data = [idx, dim, name, current_val, db_default, value_range, data_type, edit_level, core_func]

        row_fill = dim_fills.get(dim, None)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx in (1, 2, 7, 8):
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
            if row_fill:
                cell.fill = row_fill

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:I{len(params)+1}"

    # 添加汇总Sheet
    ws2 = wb.create_sheet("维度汇总")
    summary_headers = ["维度", "参数数量", "颜色标识"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    summaries = [
        ("CPU调度", 25, "蓝色底色"),
        ("内存管理", 25, "绿色底色"),
        ("磁盘IO", 25, "橙色底色"),
        ("SQL执行", 25, "红色底色"),
    ]
    for idx, (dim, count, color) in enumerate(summaries, 2):
        ws2.cell(row=idx, column=1, value=dim).font = cell_font
        ws2.cell(row=idx, column=2, value=count).font = cell_font
        ws2.cell(row=idx, column=3, value=color).font = cell_font
        fill = dim_fills.get(dim)
        if fill:
            ws2.cell(row=idx, column=1).fill = fill
        for col in range(1, 4):
            ws2.cell(row=idx, column=col).border = thin_border
            ws2.cell(row=idx, column=col).alignment = center_alignment

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 15

    wb.save(output_path)
    print(f"✓ Excel已导出: {output_path}")


def main():
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "OceanBase_v4.5_关键性能参数_100.xlsx")

    print("=" * 60)
    print("OceanBase v4.5 关键性能参数导出")
    print("=" * 60)

    # 获取数据库当前参数值
    print("\n[1/3] 从数据库获取当前参数值...")
    param_map = fetch_current_values()
    if param_map:
        print(f"  ✓ 已获取 {len(param_map)} 个参数的当前值")
    else:
        print("  ⚠ 未连接数据库，将使用默认值")

    # 校验参数是否存在
    print("\n[2/3] 校验参数列表...")
    missing = []
    for name, dim, _, _, _ in KEY_PARAMS:
        if name not in param_map:
            missing.append(name)
    if missing:
        print(f"  ⚠ 以下 {len(missing)} 个参数在当前版本中未找到:")
        for m in missing:
            print(f"    - {m}")
    else:
        print("  ✓ 所有参数均已校验通过")

    # 导出Excel
    print("\n[3/3] 导出Excel...")
    export_to_excel(KEY_PARAMS, param_map, output_path)

    # 统计
    dims = {}
    for _, dim, _, _, _ in KEY_PARAMS:
        dims[dim] = dims.get(dim, 0) + 1

    print("\n" + "=" * 60)
    print("导出完成！参数统计:")
    print("=" * 60)
    for dim, count in dims.items():
        print(f"  {dim}: {count} 个参数")
    print(f"  总计: {sum(dims.values())} 个参数")
    print(f"\n文件位置: {output_path}")


if __name__ == "__main__":
    main()
